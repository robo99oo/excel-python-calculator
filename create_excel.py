import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Calculator"

ws["A2"] = "x"
ws["A3"] = "y"
ws["A4"] = "operation"
ws["A6"] = "result"
ws["A7"] = "status"

wb.save("Calculator.xlsx")
print("Done!")